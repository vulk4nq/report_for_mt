import os
try:
    from openpyxl import Workbook, load_workbook
except:
    os.system("pip install openpyxl")
try:
    import psycopg2
except:
    os.system("pip install psycopg2")
import time
from datetime import datetime
current_time = datetime.now()


#time in minutes
def get_xlsx(minutes):
    print(minutes==5)
    db = psycopg2.connect(user="postgres",
                          # пароль, который указали при установке PostgreSQL
                          password="",
                          host="localhost",#80.89.238.221
                          port="5432",
                          database="bot")#websockets_data
    cursor = db.cursor()
   
    query = f"""SELECT max(time) FROM data;
	"""
    cursor.execute(query)
    time_ = cursor.fetchone()[0]
    if minutes != 1:
        time_ = round(time.time(),3)*1000
    
        

    query = f"""SELECT coin,tp_or_sl,change,tp,sl,distance,type_of_,time,id FROM data WHERE time between {time_-(minutes*60000)+5000} AND {time_}
	"""
    array_of_dict = {}

    cursor.execute(query)
    while True:
        # consume result over a series of iterations
        # with each iteration fetching 2000 records
        records = cursor.fetchmany(size=2000)

        if not records:
            break

        for r in records:
            
            if f"{r[0]}_{r[5]}_{r[3]}_{r[4]}_{r[6]}" in array_of_dict:

                if r[1] == 'None':
                    array_of_dict[f"{r[0]}_{r[5]}_{r[3]}_{r[4]}_{r[6]}"]['count_all'] += 1
                    array_of_dict[f"{r[0]}_{r[5]}_{r[3]}_{r[4]}_{r[6]}"]['count_minus'] += 0
                    array_of_dict[f"{r[0]}_{r[5]}_{r[3]}_{r[4]}_{r[6]}"]['coeff'] -= r[4]

                elif r[1] == 'sl':
                    array_of_dict[f"{r[0]}_{r[5]}_{r[3]}_{r[4]}_{r[6]}"]['count_all'] += 1
                    array_of_dict[f"{r[0]}_{r[5]}_{r[3]}_{r[4]}_{r[6]}"]['count_minus'] += 1
                    array_of_dict[f"{r[0]}_{r[5]}_{r[3]}_{r[4]}_{r[6]}"]['coeff'] -= r[4]

                elif r[1] == 'tp':
                    array_of_dict[f"{r[0]}_{r[5]}_{r[3]}_{r[4]}_{r[6]}"]['count_all'] += 1
                    array_of_dict[f"{r[0]}_{r[5]}_{r[3]}_{r[4]}_{r[6]}"]['count_plus'] += 1
                    array_of_dict[f"{r[0]}_{r[5]}_{r[3]}_{r[4]}_{r[6]}"]['coeff'] += r[3]
            else:
                if r[1] == 'None':
                    array_of_dict[f"{r[0]}_{r[5]}_{r[3]}_{r[4]}_{r[6]}"] = {
                        'count_all': 1,
                        'count_plus': 0,
                        'count_minus': 0,
                        'coeff': -r[4]
                    }

                elif r[1] == 'sl':
                    array_of_dict[f"{r[0]}_{r[5]}_{r[3]}_{r[4]}_{r[6]}"] = {
                        'count_all': 1,
                        'count_plus': 0,
                        'count_minus': 1,
                        'coeff': -r[4]
                    }

                elif r[1] == 'tp':
                    array_of_dict[f"{r[0]}_{r[5]}_{r[3]}_{r[4]}_{r[6]}"] = {
                        'count_all': 1,
                        'count_plus': 1,
                        'count_minus': 0,
                        'coeff': r[3]
                    }

    

    wb = Workbook()
    ws = wb.active
    
    
    count = 2
    ws[f'A1'] = 'Эффективность'  # эффективность
    ws[f'B1'] = 'Пара'  # пара
    ws[f'C1'] = 'Кол-во сделок'  # кол-во сделок
    ws[f'D1'] = 'Кол-во сделок в плюс'  # кол-во сделок+
    ws[f'E1'] = 'Кол-во сделок в минус'  # кол-во сделок-
    ws[f'F1'] = 'Side'  # side
    ws[f'G1'] = 'Distance'  # distance
    ws[f'H1'] = 'Take Profit'  # tp
    ws[f'I1'] = 'Stop Loss'  # sl
    ws[f'J1'] = 'Копировать#'
    for key in array_of_dict:
        x = key.split('_')
        ws[f'A{count}'] = array_of_dict.get(key).get('coeff')  # эффективность
        ws[f'B{count}'] = x[0]  # пара
        ws[f'C{count}'] = array_of_dict.get(
            key).get('count_all')  # кол-во сделок
        ws[f'D{count}'] = array_of_dict.get(
            key).get('count_plus')  # кол-во сделок+
        ws[f'E{count}'] = array_of_dict.get(
            key).get('count_minus')  # кол-во сделок-
        ws[f'F{count}'] = x[-1]  # side
        ws[f'G{count}'] = x[1]  # distance
        ws[f'H{count}'] = x[2]  # tp
        ws[f'I{count}'] = x[3]  # sl
        ws[f'J{count}'] = f"""
###START###
algorithmName=0=Shots Group;
info=23={current_time.day}.{current_time.month}.{current_time.year}-{x[0]} / DIS {x[1]} / TP {x[2]} / SL {x[3]};
autoStart=11=False;autoRestart=11=True;
shotRestartDelay=2=30;
stopIfTradeLatencyGreaterThan=4=1;
exchangeType=7=1;
marketType=8=3;
quoteAssets=25=;
whiteList=24={x[0]},;
blackList=24=;
useCoinDelta=11=False;
coinDeltaTimeframe=17=3600;
            """+"""
coinMinMaxDelta=18={
    "min": 0.0,
    "max": 3.0
};
useQAV24=11=False;
qav24MinMax=19={
    "min": 0.0,
    "max": 0.0
};
useBtcDelta=11=False;
btcDeltaTimeframe=17=60;
btcMinMaxDelta=18={
    "min": 0.0,
    "max": 0.0
};"""+f"""
distance=4={x[1]};
buffer=4=0.1;
followPriceDelay=2=1;
useTickerSource=20=False;
side=10=-1;
clientOrderType=14=1;
stopPricePercentage=4=3;
orderSizeUSDT=5=15;
takeProfitIsOn=11=True;
takeProfitPercentage=4={x[2]};
takeProfitOrderType=21=1;
takeProfitStatus=13=2;
useAutoPriceDown=11=False;
autoPriceDownTimer=4=0;
autoPriceDownStep=4=0;
autoPriceDownLimit=4=0;
stopLossIsOn=11=True;
stopLossPercentage=4={x[3]};
stopLossSpread=4=1;
stopLossDelay=4=0;
stopLossOrderType=22=2;
stopLossStatus=13=1;
stopLossTrailingIsOn=11=False;
stopLossTrailingSpread=4=0;

			"""  # copy
        
        count += 1
    
    if minutes == 1:
        wb.save("minute.xlsx")
    elif minutes == 5:
        wb.save("minfive.xlsx")
    elif minutes == 60:
        wb.save("hour.xlsx")
    elif minutes == 1440:
       wb.save("day.xlsx")
    elif minutes == 10080:
        wb.save("week.xlsx")
if __name__ == '__main__':
    print("starting write to excel")
    get_xlsx(1)
    print("Записал минутный")  
    get_xlsx(5)
    print("Записал 5-ти минутный")
    get_xlsx(60)
    print("Записал часовой")
    get_xlsx(1440)
    print("Записал дневной")
    get_xlsx(10080)
    print("Записал недельный")
    print("ending")
